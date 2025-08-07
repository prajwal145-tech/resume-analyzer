import os
import re
import pandas as pd
import docx2txt
import fitz  # PyMuPDF
import openpyxl
from openpyxl.styles import Font
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity
from sentence_transformers import SentenceTransformer
import spacy

nlp = spacy.load("en_core_web_sm")
model = SentenceTransformer('all-MiniLM-L6-v2')


def extract_text_from_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    if ext == '.pdf':
        text = ''
        with fitz.open(file_path) as doc:
            for page in doc:
                text += page.get_text()
        return text
    elif ext in ['.doc', '.docx']:
        return docx2txt.process(file_path)
    elif ext == '.txt':
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            return f.read()
    return ''


def extract_keywords_from_jd(jd_text, top_n=25):
    doc = nlp(jd_text)
    keywords = set()
    for token in doc:
        if token.is_alpha and not token.is_stop and token.pos_ in ['NOUN', 'PROPN', 'ADJ']:
            keywords.add(token.lemma_.lower())
    return sorted(keywords)[:top_n]


def analyze_resumes(jd_text, resume_paths, selected_keywords, top_n):
    # Remove previous result file if exists
    result_path = os.path.join('uploads', 'Analysis_Result.xlsx')
    if os.path.exists(result_path):
        os.remove(result_path)

    jd_vector = model.encode([jd_text])[0]
    results = []

    for path in resume_paths:
        filename = os.path.basename(path)
        text = extract_text_from_file(path)
        resume_vector = model.encode([text])[0]

        skill_matches = sum(1 for kw in selected_keywords if kw.lower() in text.lower())
        skill_match_percent = round(100 * skill_matches / len(selected_keywords), 2) if selected_keywords else 0
        jd_match_percent = round(100 * cosine_similarity([jd_vector], [resume_vector])[0][0], 2)

        results.append({
            'Resume': filename,
            'Skill Match %': skill_match_percent,
            'JD Match %': jd_match_percent
        })

    df = pd.DataFrame(results)

    df['Average'] = df[['Skill Match %', 'JD Match %']].mean(axis=1)
    df['Rank'] = df['Average'].rank(method='min', ascending=False).astype(int)
    df.drop(columns=['Average'], inplace=True)
    df.sort_values('Rank', inplace=True)

    df.to_excel(result_path, index=False)
    format_excel_file(result_path, top_n)
    return df, result_path


def format_excel_file(file_path, top_n):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    ws.freeze_panes = ws['A2']

    # Auto column width
    for col in ws.columns:
        max_len = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_len + 2

    # Bold top N rows
    for row in ws.iter_rows(min_row=2, max_row=top_n + 1):
        for cell in row:
            cell.font = Font(bold=True)

    # Color scale for % columns
    for col in ['B', 'C']:
        rule = ColorScaleRule(start_type='min', start_color='FFAAAA',
                              mid_type='percentile', mid_value=50, mid_color='FFFFAA',
                              end_type='max', end_color='AAFFAA')
        ws.conditional_formatting.add(f'{col}2:{col}{ws.max_row}', rule)

    wb.save(file_path)
